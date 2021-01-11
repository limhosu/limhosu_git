from openpyxl import load_workbook
import datetime
from datetime import date
from datetime import timedelta
import pandas as pd
import numpy as np
import csv
import os
os.remove("d:/파이썬/test.csv")

wb = load_workbook("d:\파이썬\sample.xlsx", data_only = True)
ws = wb.active # 현재 활성화된 sheet 가져옴

# x = input('날짜(2020-10-01): ') # 년월일을 문자형식으로 입력받기
x='2020-12'
# x=datetime.datetime.strptime(x, '%Y-%m-%d') # 입력받은 문자형 년월일을 datetime형식으로 변환
# y=x.date()
# y=x.strftime('%Y-%m')
# print(y)

for row in ws.iter_rows(min_row=2):
    b=row[0].value
    b=b.date()
    b=b.strftime('%Y-%m')
    c=(b, row[1].value, row[2].value)
    # print(c)
    for i in c:
        if i == x:
            e=(b, row[1].value, row[2].value)
            e=list(e)
            print(e)

            f=open('d:/파이썬/test.csv', 'a', encoding='utf-8', newline='')
            wr = csv.writer(f)
            wr.writerows([e])
            f.close()
# 반복문을 돌며 튜플에 있는 값을 워크시트에 값을 쓴다.

            
            # print(b, sum(row[1].value), sum(row[2].value))
wb.save("d:\파이썬\sample_formula.xlsx")